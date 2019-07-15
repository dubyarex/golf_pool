#! C:\Anaconda3\envs\py3\python

import requests, json, sys, os, openpyxl, get_picks, time, datetime
import pandas as pd
from pprint import pprint
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, Protection



    ###### For Testing Purposes ######
testing = False

if testing:
    print('Testing = True')
    tname = 'Testing ' + tname

    ###### #################### ######


    ###### User given variales ######

### Have user choose which tournament to update
# tournament_name = input('Enter Desired Tournament: ')

### Store pool excel workbook here
excel_folder = 'C:\\Stuff\\Pools\\Golf\\'
### penalty score for WD, CUT, MDF
penalty_score = 78
### Placeholder while testing code
tid = '100'  # The Open Championship  'tid' for PGA.com
# PGA.com URL given a tournament ID - 'tid'
url = f'https://statdata.pgatour.com/r/{tid}/leaderboard-v2mini.json'

live_tab = 'Live'
details_tab = 'Details'
raw_data_tab = 'Raw Data'

template_filename = 'Leaderboard_Template.xlsx'

    ###### ################### ######


    ###### Get the data ######

res = requests.get(url)
### parses json code into python data format (Dictionaries and Lists)
data = json.loads(res.text)

leaderboard = data['leaderboard']

    ######## Tournament Details ########

### List of desired tournament detail fields
# print(leaderboard.key())  # uncomment To see all possible fields
leaderboard_headers = ['tournament_id', 'tournament_name', 'start_date',
                       'end_date', 'is_started', 'is_finished', 'current_round',
                       'round_state']
                       ### round_state -- 'Official', 'Groupings Official', '--Active??--'

### Sub-Headers to desired tournament detail fields
### leaderboard['courses'][0]
course_sub_headers = ['course_id', 'course_name', 'par_in', 'par_out', 'par_total']
### leaderboard['cut_line']
cutline_sub_headers = ['cut_count', 'cut_line_score']
### combine all headers and sub-headers

### Create simplified Dictionary of desired tournament details
tournament_details = {}
### Add leaderboard_headers
for i in leaderboard_headers:
    tournament_details[i] = leaderboard[i]
### Add course_sub_headers
for i in course_sub_headers:
    tournament_details[i] = leaderboard['courses'][0][i]
### Add cutline_sub_headers
for i in cutline_sub_headers:
    tournament_details[i] = leaderboard['cut_line'][i]
### Add user define values
tournament_details['penalty_score'] = penalty_score
tournament_details['update_time'] = datetime.datetime.strptime(data['last_updated'], '%Y-%m-%dT%H:%M:%S')

    ######## ################## ########


    ###### Tournament Data Derived variables ######

tournament_headers = tournament_details.keys()

tname = tournament_details['tournament_name']
tyear = tournament_details['start_date'][:4]
excel_filename = f'{excel_folder}{tname} -- {tyear}.xlsx'

tournament_details['tournament_key'] = f'{tname} -- {tyear}'

    ###### Tournament Data Derived variables ######


    ###### Picks ######

### pick file is tournament and year specfic
picks_filename = f'{tname} {tyear} - Picks.csv'
picks_file = f'{excel_folder}{picks_filename}'

pick_data = get_picks.from_csv(picks_file)

picks = []
for row, values in enumerate(pick_data):
    if row > 0:
        pool = {}
        for i, column in enumerate(values):
            pool['_'.join(pick_data[0][i].split(' '))] = column
            # if i == 0:
            #     pool[pick_data[0][i]] = column
            # else:
            #     # short_name = column[0] + '. ' + ' '.join(column.split()[1:],)
            #     # pool[pick_data[0][i]] = short_name

        picks.append(pool)

    ###### ##### ######


    ######## Check for Withdraws ########

### Check if file exist. If yes find any 'wd#' statuses and add to prior_wd
###   where prior_wd['short_name'] = 'wd' + current_round
prior_wd = {}

if os.path.isfile(excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb[raw_data_tab]
    for header in sheet[1]:
        if header.value == 'status':
            if type(header.column) is int:
                status_col = get_column_letter(header.column)
            else:
                status_col = header.column
    for status in sheet[status_col]:
        if status.value[:2] == 'wd':
            pwd = sheet['A' + str(status.row)].value
            print(f'{pwd} withdrew from tournament. Double check scores')
            prior_wd[pwd] = status.value
    wb.close()

print('List of Prior Withdraws:')
print(prior_wd)
    ######## ################## ########


    ######## Player Details ########

### List of desired player detail fields
player_headers = ['current_position', 'start_position', 'status', 'thru',
                  'current_round', 'course_hole', 'today', 'total',
                  'total_strokes']
### leaderboard['players'][i]['player_bio']
bio_sub_headers = ['first_name', 'short_name', 'last_name']
### leaderboard['players'][i]['rounds']
rounds_sub_headers = [['r1_strokes', 'r1_tee_time'],
                      ['r2_strokes', 'r2_tee_time'],
                      ['r3_strokes', 'r3_tee_time'],
                      ['r4_strokes', 'r4_tee_time']]

### Create simplified list of Dictionary of player details
player_list = []
### create custom 'short_name' == short_name + '. ' + last_name
for player in leaderboard['players']:
    player_details = {}
    for i in player_headers:
        player_details[i] = player[i]
    for i in bio_sub_headers:
        player_details[i] = player['player_bio'][i]
    count = 0
    for i in rounds_sub_headers:
        player_details[i[0]] = player['rounds'][count]['strokes']
        player_details[i[1]] = player['rounds'][count]['tee_time']
        count += 1
    ### create custom 'short_name' == short_name + '. ' + last_name
    player_details['short_name'] = player['player_bio']['short_name'] + '. ' + player['player_bio']['last_name']
    player_details['name'] = player['player_bio']['first_name'] + ' ' + player['player_bio']['last_name']

    ### create 'adj_status'
    ### Check for previous 'wd#'
    if player_details['name'] in prior_wd.keys():
        player_details['status'] = prior_wd[player_details['name']]
    ### Identify NEW 'wd's
    if player_details['status'] == 'wd':
        player_details['status'] = player_details['status'] + str(tournament_details['current_round'])
    else:
        player_details['status'] = player_details['status']
    ### Alternate to the above:
    ### Adjust 'status' of 'wd' Look to drop the 'adj_status' values and just change 'status' in place for 'wd's


    ### append to player_list as a dictionary
    player_list.append(player_details)

            ######## Score Adjustments ########

    ### Could add code to adjust scores for players with status = ['wd', 'cut', 'mdf']
    round_scores = ['r4_strokes', 'r3_strokes', 'r2_strokes', 'r1_strokes']
    for player in player_list:
        ### adjust scores for players who Withdraw
        if player['status'][:2] == 'wd':
            for score in round_scores:
                ### If player has withdrawn, all None values = penalty score
                ### All unstarted rounds would be None so set to penalty
                if player[score] == None:
                    player[score] = penalty_score
                ### If Rd# is = wd# --
                ### r#_strokes = Higher of(r#_strokes or penalty)score
                ### That is to say if they took more strokes on that day than
                ### the penalty score, they keep their higher score
                if int(score[1]) == int(player['status'][-1]):

                    player[score] = max(player[score], penalty_score)
                ### Else (Rd# < wd#), keep score unchanged
            player['total_strokes'] = sum(player[score] for score in round_scores)
            player['total'] = player['total_strokes'] - (4 * int(tournament_details['par_total']))
            player['today'] = penalty_score - int(tournament_details['par_total'])
        ### adjust scoares for Cut players
        if player['status'] == 'cut':
            player['r3_strokes'] = penalty_score
            player['r4_strokes'] = penalty_score
            player['total_strokes'] = sum(player[score] for score in round_scores)
            player['total'] = player['total_strokes'] - (4 * int(tournament_details['par_total']))
            player['today'] = penalty_score - int(tournament_details['par_total'])
        ### adjust scores for Made Cut, DNF
        if player['status'] == 'mdf':
            player['r4_strokes'] = penalty_score
            player['total_strokes'] = sum(player[score] for score in round_scores)
            player['total'] = player['total_strokes'] - (4 * int(tournament_details['par_total']))
            player['today'] = penalty_score - int(tournament_details['par_total'])

            ######## ################# ########

    ######## ############## ########

    ######## Pools Total Scores ########


pick_num = (list(picks[0].keys()))
pick_num.pop(0)
for pool in picks:
    total_value = 0
    for pick in pick_num:
        for player in player_list:
            if player['short_name'] == pool[pick]:
                total_value += player['total']
                ### Add Players Scores next to name
                if player['total'] > 0:
                    pool[pick] += f' (+{player["total"]})'
                elif player['total'] == 0:
                    pool[pick] += f' (E)'
                else:
                    pool[pick] += f' ({player["total"]})'
    pool['Total_Score'] = total_value

picks = sorted(picks, key = lambda pool: pool['Total_Score'])


player_columns = player_list[0].keys()

raw_data_columns = ['name',
                    'short_name',
                    'current_position',
                    'total',
                    'thru',
                    'today',
                    'r1_strokes',
                    'r2_strokes',
                    'r3_strokes',
                    'r4_strokes',
                    'total_strokes',
                    'current_round',
                    'status',
                    'course_hole',
                    'r1_tee_time',
                    'r2_tee_time',
                    'r3_tee_time',
                    'r4_tee_time']

    ######## ############## ########



### Check if file and sheet exist. If no, create file and/or sheet as needed
if os.path.isfile(excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
    ### Check if sheets exists
    if raw_data_tab not in wb.sheetnames:
        wb.create_sheet(raw_data_tab)
    if details_tab not in wb.sheetnames:
        wb.create_sheet(details_tab)

### Create a new excel file
else:
    wb = openpyxl.load_workbook(template_filename)


### Add tournament details to Details sheet
sheet = wb[details_tab]

row_num = 1
for head in tournament_headers:
    sheet.cell(row=row_num, column=1).value = head
    sheet.cell(row=row_num, column=2).value = tournament_details[head]
    row_num += 1

### Add player data to Raw Data sheet
sheet = wb[raw_data_tab]

### Header Row
for col, header in enumerate(raw_data_columns):
    sheet.cell(row=1, column=(col + 1)).value = header
### Player Data
for row, player in enumerate(player_list):
    for col, header in enumerate(raw_data_columns):
        sheet.cell(row=(row+2), column=(col+1)).value = player[header]

### Update Live Tab
live_table = []
sheet = wb[live_tab]
live_row = []
for col, header in enumerate(picks[0].keys()):
    if header != 'Total_Score':
        sheet.cell(row=3, column=(col + 1)).value = header.replace("_", " ")
        live_row.append(header)
    else:
        sheet.cell(row=3, column=(col + 2)).value = header.replace("_", " ")
        live_row.append(header)
live_table.append(live_row)

for row, name in enumerate(picks):
    live_row = []
    for col, header in enumerate(picks[0].keys()):
        if header != 'Total_Score':
            sheet.cell(row=(row+4), column=(col+1)).value = name[header]
            live_row.append(name[header])
        else:
            sheet.cell(row=(row+4), column=(col+2)).value = name[header]
            live_row.append(name[header])
    live_table.append(live_row)
### Add update time to Live sheet, cell 'J3'
sheet.cell(row=(3), column=(10)).value = tournament_details['update_time']

### Find Sheet Index of Live Tab and make it the active sheet
for i, sht_name in enumerate(wb.sheetnames):
    if sht_name == live_tab:
        wb.active = i

### Set all other sheets to not-active
for sheet in wb.sheetnames:
    if sheet != live_tab:
        wb[sheet].sheet_view.tabSelected = False



wb.save(excel_filename)
print(f'\nData written to: {excel_filename}\n')
print(f'Scores updated as of -- {tournament_details["update_time"]}')
print(f'Script run at -- {time.ctime()}')

# pprint(picks[0])
# pprint(tournament_details)
# pprint(player_list[0])

tournaments = []
tournament_details['players'] = player_list
tournament_details['picks'] = picks
tournaments.append(tournament_details)

# with open('tournaments.json', 'w') as fout:
#     json.dump(tournaments, fout, default=str)

# pprint(tournaments)

'''
### Get Dictionary for Desired Tournament
for leaderboard in data['Leaderboards']:
    if leaderboard['Tournament'] == tournament_name:
        tournament_data = leaderboard

player_data = tournament_data['Players']

player_list = []
for player in player_data:
    player_list.append(player['Name'])

player_count = len(player_list)
column_titles = ['Name', 'Short Name', 'CurrentPosition', 'Total', 'After', 'Today', 'Round 1', 'Round 2', 'Round 3', 'Round 4', 'TotalStrokes']
col_count = len(column_titles)


### Check for existing file and open or create new file
if os.path.isfile('leader_test_2019.xlsx'):
    wb = openpyxl.load_workbook('leader_test_2019.xlsx')
### Create a new excel file
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Raw Data'

### Select Raw Data Sheet
sheet = wb['Raw Data']

### Creating Title for Data Table
# sheet.merge_cells('A1:' + get_column_letter(col_count) + '1')
sheet['A1'] = tournament_name
sheet['A1'].font = Font(size=24, bold=True, underline="single")
sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

### Adds Columns Titles to Data Table
for i in range(1, len(column_titles) + 1):
    sheet.cell(row=2, column=i).value = column_titles[i-1]
    sheet.cell(row=2, column=i).font = Font(bold=True)

### Function to Generate Short Names
def shorten_name(full_name):
    names = full_name.split()
    name_count = len(names)
    short_name = names[0][0] + '. ' + ' '.join(names[1:])
    return(short_name)

### Add Short_Names and separate Round Scores
for player in player_data:
    player['Short Name'] = shorten_name(player['Name'])
    for round_num in list(enumerate(['Round 1', 'Round 2', 'Round 3', 'Round 4'])):
        player[round_num[1]] = player['Rounds'][round_num[0]]

### Iterate through players and write data to excel sheet
row_num = 3
for player in player_data:
    col_num = 1
    for col in column_titles:
        sheet.cell(row=row_num, column=col_num).value = player[col]
        col_num += 1
    row_num += 1



# col_num = 1
# for col in parser.keys():               # Iterate thorugh each Column
#   for player in range(player_count):
#       if col_num > 2:
#           if leaderboard.select(parser[col])[player].getText() == 'E':
#               sheet.cell(row=player+3, column=col_num).value = 0
#           elif leaderboard.select(parser[col])[player].getText() == '-' or leaderboard.select(parser[col])[player].getText() == '--':
#               sheet.cell(row=player+3, column=col_num).value = 'Cut'
#           else:
#               sheet.cell(row=player+3, column=col_num).value = int(leaderboard.select(parser[col])[player].getText())  # Iterate through each row in a Column
#       else:
#           sheet.cell(row=player+3, column=col_num).value = leaderboard.select(parser[col])[player].getText()
#   col_num += 1


sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 20

wb.save('leader_test_2019.xlsx')

pprint(player_data[70])
'''
