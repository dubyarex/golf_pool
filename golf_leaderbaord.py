#! C:\Anaconda3\envs\py3\python

import requests, json, sys, os, openpyxl
from pprint import pprint
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, Protection

### Store pool excel workbook here
excel_folder = 'C:\\Stuff\\pools\\Golf\\'


### Have user choose which tournament to update
# tournament_name = input('Enter Desired Tournament: ')

### Placeholder while testing code
tid = '014'  # The Masters  'tid' for PGA.com
# PGA.com URL given a tournament ID - 'tid'
url = 'https://statdata.pgatour.com/r/{}/leaderboard-v2mini.json'.format(tid)
res = requests.get(url)

### parses json code into python data format (Dictionaries and Lists)
data = json.loads(res.text)

leaderboard = data['leaderboard']

   ######## Tournament Detials ########

### List of desired tournament detail fields
# print(leaderboard.key())  # To see all possible fields
leaderboard_headers = ['tournament_id', 'tournament_name', 'start_date', 
                       'end_date', 'is_started', 'is_finished', 'current_round', 
                       'round_state']
                       ### round_state -- 'Official', 'Groupings Official', '--Active??--'

### Sub-Headers to desired tournament detail fields
### leaderboard['courses'][0]
course_sub_headers = ['course_id', 'course_name', 'par_in', 'par_out', 'par_total']
### leaderboard['cut_line']
cutline_sub_headers = ['cut_count', 'cut_line_score']
### combine all headers and sub-headers

### Create simplified Dictionary of desired tournament details
tournament_details = {}
### Add leaderboard_headers
for i in leaderboard_headers:
	tournament_details[i] = leaderboard[i]
### Add course_sub_headers
for i in course_sub_headers:
	tournament_details[i] = leaderboard['courses'][0][i]
### Add cutline_sub_headers
for i in cutline_sub_headers:
	tournament_details[i] = leaderboard['cut_line'][i]

tournament_headers = tournament_details.keys()
tname = tournament_details['tournament_name']
tyear = tournament_details['start_date'][:4]

   ######## ################## ########



   ######## Player Detials ########

### List of desired player detail fields
player_headers = ['current_position', 'start_position', 'status', 'thru', 
                  'current_round', 'course_hole', 'today', 'total', 
                  'total_strokes']
### leaderboard['players'][i]['player_bio']
bio_sub_headers = ['first_name', 'short_name', 'last_name']
### leaderboard['players'][i]['rounds']
rounds_sub_headers = [['r1_strokes', 'r1_tee_time'], 
                      ['r2_strokes', 'r2_tee_time'], 
                      ['r3_strokes', 'r3_tee_time'],
                      ['r4_strokes', 'r4_tee_time']]

### Create simplified list of Dictionary of player details
player_list = []
### create custom 'short_name' == short_name + '. ' + last_name
for player in leaderboard['players']:
	player_details = {}
	for i in player_headers:
		player_details[i] = player[i]
	for i in bio_sub_headers:
		player_details[i] = player['player_bio'][i]
	count = 0
	for i in rounds_sub_headers:
		player_details[i[0]] = player['rounds'][count]['strokes']
		player_details[i[1]] = player['rounds'][count]['tee_time']
		count += 1
	### create custom 'short_name' == short_name + '. ' + last_name
	player_details['short_name'] = player['player_bio']['short_name'] + '. ' + player['player_bio']['last_name']

	player_list.append(player_details)

	### append to player_list as a dictionary


player_columns = player_list[0].keys()


   ######## ############## ########

'''
excel_filename = '{}{} -- {}.xlsx'.format(excel_folder, tname, tyear)
details_tab = 'Details'
raw_data_tab = 'Raw Data'

### Check if file and sheet exist. If no, create file and/or sheet as needed
if os.path.isfile(excel_filename):
	wb = openpyxl.load_workbook(excel_filename)
	### Check if sheets exists
	if raw_data_tab not in wb.sheetnames:
		wb.create_sheet(raw_data_tab)
	if details_tab not in wb.sheetnames:
		wb.create_sheet(details_tab)

### Create a new excel file
else:
	wb = openpyxl.Workbook()
	sheet = wb.active
	### Rename sheet
	sheet.title = raw_data_tab
	wb.create_sheet(details_tab)

sheet = wb[details_tab]

row_num = 1
for head in tournament_headers:
	sheet.cell(row=row_num, column=1).value = head
	sheet.cell(row=row_num, column=2).value = tournament_details[head]
	row_num += 1


wb.save(excel_filename)
'''

'''
### Get Dictionary for Desired Tournament
for leaderboard in data['Leaderboards']:
	if leaderboard['Tournament'] == tournament_name:
		tournament_data = leaderboard

player_data = tournament_data['Players']

player_list = []
for player in player_data:
	player_list.append(player['Name'])

player_count = len(player_list)
column_titles = ['Name', 'Short Name', 'CurrentPosition', 'Total', 'After', 'Today', 'Round 1', 'Round 2', 'Round 3', 'Round 4', 'TotalStrokes']
col_count = len(column_titles)


### Check for existing file and open or create new file
if os.path.isfile('leader_test_2019.xlsx'):
	wb = openpyxl.load_workbook('leader_test_2019.xlsx')
### Create a new excel file
else:
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = 'Raw Data'

### Select Raw Data Sheet
sheet = wb['Raw Data']

### Creating Title for Data Table
# sheet.merge_cells('A1:' + get_column_letter(col_count) + '1')
sheet['A1'] = tournament_name
sheet['A1'].font = Font(size=24, bold=True, underline="single")
sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

### Adds Columns Titles to Data Table
for i in range(1, len(column_titles) + 1):
	sheet.cell(row=2, column=i).value = column_titles[i-1]
	sheet.cell(row=2, column=i).font = Font(bold=True)

### Function to Generate Short Names
def shorten_name(full_name):
	names = full_name.split()
	name_count = len(names)
	short_name = names[0][0] + '. ' + ' '.join(names[1:])
	return(short_name)

### Add Short_Names and separate Round Scores
for player in player_data:
	player['Short Name'] = shorten_name(player['Name'])
	for round_num in list(enumerate(['Round 1', 'Round 2', 'Round 3', 'Round 4'])):
		player[round_num[1]] = player['Rounds'][round_num[0]]

### Iterate through players and write data to excel sheet
row_num = 3
for player in player_data:
	col_num = 1
	for col in column_titles:
		sheet.cell(row=row_num, column=col_num).value = player[col]
		col_num += 1
	row_num += 1



# col_num = 1
# for col in parser.keys():               # Iterate thorugh each Column
# 	for player in range(player_count):
# 		if col_num > 2:
# 			if leaderboard.select(parser[col])[player].getText() == 'E':
# 				sheet.cell(row=player+3, column=col_num).value = 0
# 			elif leaderboard.select(parser[col])[player].getText() == '-' or leaderboard.select(parser[col])[player].getText() == '--':
# 				sheet.cell(row=player+3, column=col_num).value = 'Cut'
# 			else:
# 				sheet.cell(row=player+3, column=col_num).value = int(leaderboard.select(parser[col])[player].getText())  # Iterate through each row in a Column 
# 		else:
# 			sheet.cell(row=player+3, column=col_num).value = leaderboard.select(parser[col])[player].getText()
# 	col_num += 1


sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 20

wb.save('leader_test_2019.xlsx')

pprint(player_data[70])
'''